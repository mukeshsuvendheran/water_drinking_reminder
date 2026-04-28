import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
from plyer import notification



BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = os.path.join(BASE_DIR, "drink_water_log.xlsx")

#FILE_NAME = "/EXCEL_READING/python_reminder/drink_water_log.xlsx"

if not os.path.exists(FILE_NAME):    # Create Excel file if not exists
#create, append & save
    wb = Workbook()   #Creates a new Excel workbook.
    ws = wb.active   #Accesses the active worksheet in the workbook. 
    ws.title = "Water Reminder"         #Sets the title 
    ws.append(["Time", "Reminder", "Status"])    #set the row at 1st column
    wb.save(FILE_NAME)                  #Save

print("Water Reminder Started... 💧")

while True:
    #3 variables for value 
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")     # Get current time
    reminder_message = "Drink Water 💧"                            # Reminder message
    status = "Pending"     
    #notification
    notification.notify(
        title="Water Reminder 💧",
        message=f"It's {current_time}\nTime to drink water!",
        timeout=10)
   #append & save
    wb = load_workbook(FILE_NAME)   #Open existing file
    ws = wb.active                                       #Access active sheet
    ws.append([current_time, reminder_message, status])  #append the three values in the row
    wb.save(FILE_NAME)
    time.sleep(3600)  # Wait for 1 min before next reminder